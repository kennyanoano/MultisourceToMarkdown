import openpyxl

def convert_xlsx_to_md(file_path):
    workbook = openpyxl.load_workbook(file_path)
    md_content = ""

    for sheet_name in workbook.sheetnames:
        md_content += f"## {sheet_name}\n\n"  # シート名を見出しとして追加

        sheet = workbook[sheet_name]
        first_row = True
        for row in sheet.iter_rows(values_only=True):
            # Markdownのテーブル形式に変換
            md_row = "| " + " | ".join([str(cell) if cell is not None else "" for cell in row]) + " |"
            md_content += md_row + "\n"

            # ヘッダー行の下に区切り行を追加
            if first_row:
                separators = "| " + " | ".join(["---"] * len(row)) + " |"
                md_content += separators + "\n"
                first_row = False

        md_content += "\n"  # シート間に空行を追加

    return md_content